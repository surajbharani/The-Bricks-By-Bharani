"""Excel read/write via openpyxl. Clear error if library is missing."""


def _openpyxl():
    try:
        import openpyxl
        return openpyxl
    except ImportError:
        raise RuntimeError("Excel support requires openpyxl. Run: pip install openpyxl")


def read(path, sheet=0):
    """Read an Excel file and return a list of dicts (first row = headers).
    sheet can be an index (int) or sheet name (str)."""
    ox = _openpyxl()
    wb = ox.load_workbook(str(path), read_only=True, data_only=True)
    if isinstance(sheet, str):
        ws = wb[sheet]
    else:
        ws = list(wb.worksheets)[sheet]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return []
    headers = [str(c) if c is not None else f"col{i}" for i, c in enumerate(rows[0])]
    return [dict(zip(headers, r)) for r in rows[1:]]


def read_all_sheets(path):
    """Read all sheets from an Excel file → {sheet_name: [list of dicts]}."""
    ox = _openpyxl()
    wb = ox.load_workbook(str(path), read_only=True, data_only=True)
    result = {}
    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            result[ws.title] = []
            continue
        headers = [str(c) if c is not None else f"col{i}" for i, c in enumerate(rows[0])]
        result[ws.title] = [dict(zip(headers, r)) for r in rows[1:]]
    wb.close()
    return result


def write(path, rows, sheet_name="Sheet1"):
    """Write a list of dicts to an Excel file (first row = header keys)."""
    ox = _openpyxl()
    wb = ox.Workbook()
    ws = wb.active
    ws.title = sheet_name
    if not rows:
        wb.save(str(path))
        return
    headers = list(rows[0].keys())
    ws.append(headers)
    for r in rows:
        ws.append([r.get(h) for h in headers])
    wb.save(str(path))


def write_multi_sheet(path, sheets):
    """Write multiple sheets: sheets = {sheet_name: [list of dicts]}."""
    ox = _openpyxl()
    wb = ox.Workbook()
    first = True
    for name, rows in sheets.items():
        ws = wb.active if first else wb.create_sheet(name)
        ws.title = name
        first = False
        if not rows:
            continue
        headers = list(rows[0].keys())
        ws.append(headers)
        for r in rows:
            ws.append([r.get(h) for h in headers])
    wb.save(str(path))
